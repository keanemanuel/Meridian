"""XLSX export: room, applicant and panel views in one workbook (FR-53).

Clashes are marked red (FR-54): a filled cell plus a coloured, bold font, so
the flag survives both screen viewing and black-and-white printing. Amber
capacity warnings on the conflicts sheet get the equivalent amber treatment.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from iff_scheduler.domain.enums import Severity
from iff_scheduler.domain.models import Conflict
from iff_scheduler.export.applicant_view import ApplicantViewRow
from iff_scheduler.export.panel_view import PanelView
from iff_scheduler.export.room_view import RoomView

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
AMBER_FONT = Font(color="9C6500")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)
LOCKED_FONT = Font(italic=True)

_INVALID_SHEET_CHARS = set("[]:*?/\\")
_MAX_SHEET_NAME = 31


def _sheet_name(raw: str) -> str:
    """Excel sheet names: <=31 chars, no `[]:*?/\\` (FR-50, FR-52 — one sheet
    per room/day or per panel, so names must survive real ids and labels)."""
    cleaned = "".join(c for c in raw if c not in _INVALID_SHEET_CHARS).strip()
    return (cleaned or "Sheet")[:_MAX_SHEET_NAME]


def _unique_sheet_name(wb: Workbook, raw: str) -> str:
    base = _sheet_name(raw)
    existing = {ws.title for ws in wb.worksheets}
    if base not in existing:
        return base
    n = 1
    while True:
        suffix = f" ({n})"
        candidate = base[: _MAX_SHEET_NAME - len(suffix)] + suffix
        if candidate not in existing:
            return candidate
        n += 1


def _autosize(ws: Worksheet, max_width: int = 40) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            assert col is not None
            widths[col] = max(widths.get(col, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, width + 2)


def _header_row(ws: Worksheet, values: list[str]) -> None:
    ws.append(values)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _write_room_sheet(wb: Workbook, view: RoomView) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, f"{view.room_id} {view.day_label}"))
    _header_row(
        ws,
        ["Slot"] + [f"{pid} ({view.panel_divisions[pid].value})" for pid in view.panel_ids],
    )
    for row in view.rows:
        line = [f"{row.start_time.strftime('%H:%M')}-{row.end_time.strftime('%H:%M')}"]
        for pid in view.panel_ids:
            cell = row.cells[pid]
            line.append(f"{cell.full_name} ({cell.sub_division})" if cell is not None else "")
        ws.append(line)

    for r_idx, row in enumerate(view.rows, start=2):
        for c_idx, pid in enumerate(view.panel_ids, start=2):
            cell = row.cells[pid]
            if cell is None:
                continue
            xl_cell = ws.cell(row=r_idx, column=c_idx)
            if cell.is_clash:
                xl_cell.fill = RED_FILL
                xl_cell.font = RED_FONT
            elif cell.is_locked:
                xl_cell.font = LOCKED_FONT
    _autosize(ws)


APPLICANT_HEADER = [
    "Applicant ID",
    "Full name",
    "Email",
    "Choice 1 division",
    "Choice 1 sub-division",
    "Choice 1 panel",
    "Choice 1 room",
    "Choice 1 date",
    "Choice 1 start",
    "Choice 1 end",
    "Choice 2 division",
    "Choice 2 sub-division",
    "Choice 2 panel",
    "Choice 2 room",
    "Choice 2 date",
    "Choice 2 start",
    "Choice 2 end",
]
# 1-indexed column ranges for each choice's block, for clash highlighting.
_CHOICE1_COLS = range(4, 11)
_CHOICE2_COLS = range(11, 18)


def _write_applicant_sheet(wb: Workbook, rows: Sequence[ApplicantViewRow]) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, "Applicants"))
    _header_row(ws, APPLICANT_HEADER)

    for row in rows:
        ws.append(
            [
                row.applicant_id,
                row.full_name,
                row.email,
                *(
                    [
                        row.choice1.division.value,
                        row.choice1.sub_division,
                        row.choice1.panel_id,
                        row.choice1.room,
                        row.choice1.date.isoformat(),
                        row.choice1.start_time.strftime("%H:%M"),
                        row.choice1.end_time.strftime("%H:%M"),
                    ]
                    if row.choice1 is not None
                    else [""] * 7
                ),
                *(
                    [
                        row.choice2.division.value,
                        row.choice2.sub_division,
                        row.choice2.panel_id,
                        row.choice2.room,
                        row.choice2.date.isoformat(),
                        row.choice2.start_time.strftime("%H:%M"),
                        row.choice2.end_time.strftime("%H:%M"),
                    ]
                    if row.choice2 is not None
                    else [""] * 7
                ),
            ]
        )

    for r_idx, row in enumerate(rows, start=2):
        if row.choice1 is not None and row.choice1.is_clash:
            for col in _CHOICE1_COLS:
                cell = ws.cell(row=r_idx, column=col)
                cell.fill = RED_FILL
                cell.font = RED_FONT
        if row.choice2 is not None and row.choice2.is_clash:
            for col in _CHOICE2_COLS:
                cell = ws.cell(row=r_idx, column=col)
                cell.fill = RED_FILL
                cell.font = RED_FONT
    _autosize(ws)


def _write_panel_sheet(wb: Workbook, view: PanelView) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, f"Panel {view.panel_id}"))
    _header_row(ws, ["Date", "Start", "End", "Applicant ID", "Full name", "Sub-division", "Choice"])
    for row in view.rows:
        ws.append(
            [
                row.date.isoformat(),
                row.start_time.strftime("%H:%M"),
                row.end_time.strftime("%H:%M"),
                row.applicant_id,
                row.full_name,
                row.sub_division,
                row.choice_index,
            ]
        )
    for r_idx, row in enumerate(view.rows, start=2):
        if row.is_clash:
            for col in range(1, 8):
                cell = ws.cell(row=r_idx, column=col)
                cell.fill = RED_FILL
                cell.font = RED_FONT
        elif row.is_locked:
            ws.cell(row=r_idx, column=5).font = LOCKED_FONT
    _autosize(ws)


def _write_conflicts_sheet(wb: Workbook, conflicts: Sequence[Conflict]) -> None:
    ws = wb.create_sheet(_unique_sheet_name(wb, "Conflicts"))
    _header_row(ws, ["Applicant ID", "Severity", "Type", "Message"])
    for conflict in conflicts:
        ws.append([conflict.applicant_id, conflict.severity.value, conflict.type, conflict.message])
    for r_idx, conflict in enumerate(conflicts, start=2):
        fill, font = (
            (RED_FILL, RED_FONT) if conflict.severity == Severity.RED else (AMBER_FILL, AMBER_FONT)
        )
        for col in range(1, 5):
            cell = ws.cell(row=r_idx, column=col)
            cell.fill = fill
            cell.font = font
    _autosize(ws)


def write_xlsx(
    path: Path,
    room_views: Sequence[RoomView],
    applicant_rows: Sequence[ApplicantViewRow],
    panel_views: Sequence[PanelView],
    conflicts: Sequence[Conflict],
) -> None:
    """Write one workbook: a sheet per room/day, an Applicants sheet, a sheet
    per panel, and a Conflicts sheet (FR-50..FR-54)."""
    wb = Workbook()
    default_sheet = wb.active
    assert default_sheet is not None
    wb.remove(default_sheet)

    for room_view in room_views:
        _write_room_sheet(wb, room_view)
    _write_applicant_sheet(wb, applicant_rows)
    for panel_view in panel_views:
        _write_panel_sheet(wb, panel_view)
    _write_conflicts_sheet(wb, conflicts)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
